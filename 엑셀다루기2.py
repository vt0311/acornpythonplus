import pandas as pd
import numpy as np
import matplotlib.pyplot as plt


# 현재 프로그램이 작동중인 디렉토리에서 파일 읽기
#data = pd.read_excel('./weight.xlsx', header=0)

#print(data.head())

print('===================================')
# 앞의 데이터 4줄만 출력하라.
#print(data.tail())

#print('키평균:', np.mean(data.height)) 
#print('몸무게평균:', np.mean(data.weight))
#
#plt.plot(data.height, data.weight )
#plt.show()


from openpyxl import Workbook
wb = Workbook() # 새로운 문서 만들기

# grab the active worksheet
ws = wb.active  # 첫번째 활성화된 시트 가져오기

# Data can be assigned directly to cells
ws['A1'] = 42 



# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()
ws['B1'] = '평창'

# Rows can also be appended
ws.append([1, 2, 4]) # 행추가하기


# Save the file
wb.save("openpyxl_sample.xlsx")